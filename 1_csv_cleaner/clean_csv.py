"""
CSV Data Cleaner & Report Generator
------------------------------------
Reads a CSV file, cleans the data, and generates an Excel + PDF summary report.
Usage: python clean_csv.py --input data.csv --output cleaned_data
"""

import argparse
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


# ── 1. LOAD ───────────────────────────────────────────────────────────────────

def load_csv(filepath):
    print(f"[INFO] Loading: {filepath}")
    df = pd.read_csv(filepath)
    print(f"[INFO] Loaded {len(df)} rows, {len(df.columns)} columns.")
    return df


# ── 2. CLEAN ──────────────────────────────────────────────────────────────────

def clean_data(df):
    report = {}

    # Track original state
    original_rows = len(df)
    report["original_rows"] = original_rows
    report["original_cols"] = len(df.columns)

    # Remove fully empty rows
    df.dropna(how="all", inplace=True)

    # Remove duplicate rows
    dupes = df.duplicated().sum()
    df.drop_duplicates(inplace=True)
    report["duplicates_removed"] = int(dupes)

    # Strip whitespace from string columns
    for col in df.select_dtypes(include="object").columns:
        df[col] = df[col].str.strip()

    # Fix column names (lowercase, replace spaces with underscores)
    df.columns = [c.strip().lower().replace(" ", "_") for c in df.columns]

    # Convert numeric columns stored as strings
    for col in df.columns:
        try:
            converted = pd.to_numeric(df[col], errors="coerce")
            # Only convert if >70% of non-null values are numeric
            if converted.notna().sum() / max(df[col].notna().sum(), 1) > 0.7:
                df[col] = converted
        except Exception:
            pass

    # Count and fill missing values
    missing_before = df.isna().sum().sum()
    for col in df.columns:
        if df[col].dtype in ["float64", "int64"]:
            df[col].fillna(df[col].median(), inplace=True)
        else:
            df[col].fillna("UNKNOWN", inplace=True)
    report["missing_filled"] = int(missing_before)

    report["cleaned_rows"] = len(df)
    report["rows_removed"] = original_rows - len(df)

    print(f"[CLEAN] Duplicates removed : {report['duplicates_removed']}")
    print(f"[CLEAN] Missing values filled: {report['missing_filled']}")
    print(f"[CLEAN] Rows after cleaning : {report['cleaned_rows']}")

    return df, report


# ── 3. EXPORT TO EXCEL ────────────────────────────────────────────────────────

def export_excel(df, output_path):
    filepath = output_path + ".xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned Data"

    header_fill = PatternFill("solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF")

    # Write headers
    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Write data rows
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value)) for cell in col if cell.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    wb.save(filepath)
    print(f"[EXPORT] Excel saved: {filepath}")
    return filepath


# ── 4. EXPORT PDF SUMMARY ─────────────────────────────────────────────────────

def export_pdf_summary(df, report, output_path):
    filepath = output_path + "_summary.pdf"
    doc = SimpleDocTemplate(filepath, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    # Title
    story.append(Paragraph("CSV Data Cleaning Report", styles["Title"]))
    story.append(Spacer(1, 12))

    # Summary stats table
    story.append(Paragraph("Cleaning Summary", styles["Heading2"]))
    summary_data = [
        ["Metric", "Value"],
        ["Original Rows", str(report["original_rows"])],
        ["Original Columns", str(report["original_cols"])],
        ["Duplicates Removed", str(report["duplicates_removed"])],
        ["Missing Values Filled", str(report["missing_filled"])],
        ["Final Row Count", str(report["cleaned_rows"])],
        ["Rows Removed Total", str(report["rows_removed"])],
    ]
    t = Table(summary_data, colWidths=[220, 220])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1A3A5C")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F0F4F8"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (1, -1), "CENTER"),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 16))

    # Column overview
    story.append(Paragraph("Column Overview", styles["Heading2"]))
    col_data = [["Column Name", "Data Type", "Non-Null Count"]]
    for col in df.columns:
        col_data.append([col, str(df[col].dtype), str(df[col].notna().sum())])

    ct = Table(col_data, colWidths=[180, 140, 120])
    ct.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4A7AAA")),
        ("TEXTCOLOR",  (0, 0), (-1, 0), colors.white),
        ("FONTNAME",   (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F7F9FC"), colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("PADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(ct)

    doc.build(story)
    print(f"[EXPORT] PDF summary saved: {filepath}")
    return filepath


# ── MAIN ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="CSV Data Cleaner & Report Generator")
    parser.add_argument("--input",  required=True, help="Path to input CSV file")
    parser.add_argument("--output", default="cleaned_output", help="Output file name (no extension)")
    args = parser.parse_args()

    if not os.path.exists(args.input):
        print(f"[ERROR] File not found: {args.input}")
        return

    df = load_csv(args.input)
    df_clean, report = clean_data(df)
    export_excel(df_clean, args.output)
    export_pdf_summary(df_clean, report, args.output)
    print("\n✅ Done! Files saved.")


if __name__ == "__main__":
    main()
